import uuid
import csv
import json
from pathlib import Path
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from fastapi import FastAPI, UploadFile, File, BackgroundTasks, HTTPException, Request
import shutil
import jobs
import processor
import convertor
from config import GEMINI_MODEL_NAME
from jobs import init_db
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
import openpyxl
from openpyxl.styles import Font, PatternFill
from fastapi.responses import FileResponse

UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR = Path("outputs")
OUTPUT_DIR.mkdir(exist_ok=True)

app = FastAPI(title="Invoice Processor")
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)


@app.on_event("startup")
def startup():
    init_db()  # creates the SQLite table if it doesn't exist

app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
def root():
    return FileResponse("static/index.html")


# ─── SINGLE FILE ─────────────────────────────────────────────
@app.post("/process")
@limiter.limit("10/minute")
async def process_single(request: Request, file: UploadFile = File(...)):
    file_path = UPLOAD_DIR / file.filename
    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    try:
        from extractor import extract_invoice
        invoice = extract_invoice(str(file_path))
        return invoice.model_dump()
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))

@app.get("/health")
def health():
    return {"status": "ok", "model": GEMINI_MODEL_NAME}

# ─── BATCH ───────────────────────────────────────────────────
@app.post("/batch")
@limiter.limit("5/minute")
async def process_batch(
        request: Request,
        background_tasks: BackgroundTasks,
        files: list[UploadFile] = File(...)
):
    job_id = str(uuid.uuid4())
    jobs.create_job(job_id, total_files=len(files))
    batch_dir = UPLOAD_DIR / job_id
    batch_dir.mkdir()
    for file in files:
        file_path = batch_dir / file.filename
        with open(file_path, "wb") as f:
            shutil.copyfileobj(file.file, f)
    background_tasks.add_task(processor.run_batch, str(batch_dir), job_id)
    return {"job_id": job_id, "message": "Batch started"}

# ─── STATUS ──────────────────────────────────────────────────
@app.get("/status/{job_id}")
def get_status(job_id: str):
    row = jobs.get_job(job_id)
    if not row:
        raise HTTPException(status_code=404, detail="Job not found")
    return {"job_id": job_id, "status": row["status"]}


# ─── EXPORT ──────────────────────────────────────────────────
@app.get("/export/{job_id}")
def export_csv(job_id: str):
    row = jobs.get_job(job_id)
    if not row:
        raise HTTPException(404, "Job not found")
    if row["status"] != "done":
        raise HTTPException(400, "Job not finished yet")

    results = json.loads(row["results"]) if row["results"] else []
    output_path = str(OUTPUT_DIR / f"{job_id}.csv")

    # Always generate CSV even if no results
    convertor.convert_to_csv(results, [], output_path)  # pass empty errors list

    if not Path(output_path).exists():
        # fallback: create a minimal CSV with headers only
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["filename", "vendor", "total", "category", "confidence"])
            writer.writeheader()

    return FileResponse(output_path, media_type="text/csv", filename=f"invoices_{job_id}.csv")


@app.get("/results/{job_id}")
def get_results(job_id: str):
    row = jobs.get_job(job_id)
    if not row:
        raise HTTPException(404, "Job not found")
    if row["status"] != "done":
        raise HTTPException(400, "Job not finished yet")

    results = json.loads(row["results"]) if row["results"] else []
    return {"job_id": job_id, "results": results}

@app.get("/progress/{job_id}")
def get_progress(job_id: str):
    row = jobs.get_job(job_id)
    if not row:
        raise HTTPException(404, "Job not found")
    return {
        "job_id": job_id,
        "status": row["status"],
        "total": row["total_files"],
        "processed": row["processed_files"]
    }

@app.get("/export/summary/{job_id}")
def export_summary_csv(job_id: str):
    row = jobs.get_job(job_id)
    if not row:
        raise HTTPException(404, "Job not found")
    if row["status"] != "done":
        raise HTTPException(400, "Job not finished yet")

    results = json.loads(row["results"]) if row["results"] else []
    # Remove invoices that have no meaningful data
    results = [
        inv for inv in results
        if inv.get("invoice_number") or inv.get("vendor_name") or inv.get("total_amount_due")
    ]
    output_path = str(OUTPUT_DIR / f"{job_id}_summary.csv")

    # Write a simplified CSV: one row per invoice, no line items
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        fieldnames = [
            "invoice_number", "vendor_name", "invoice_date",
            "total_amount_due", "currency", "category", "confidence"
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for inv in results:
            writer.writerow({
                "invoice_number": inv.get("invoice_number", ""),
                "vendor_name": inv.get("vendor_name", ""),
                "invoice_date": inv.get("invoice_date", ""),
                "total_amount_due": inv.get("total_amount_due", ""),
                "currency": inv.get("currency", ""),
                "category": inv.get("category", ""),
                "confidence": inv.get("confidence", "")
            })

    return FileResponse(
        output_path,
        media_type="text/csv",
        filename=f"invoices_summary_{job_id}.csv"
    )

@app.get("/export/excel/{job_id}")
def export_excel(job_id: str):
    row = jobs.get_job(job_id)
    if not row:
        raise HTTPException(404, "Job not found")
    if row["status"] != "done":
        raise HTTPException(400, "Job not finished yet")

    results = json.loads(row["results"]) if row["results"] else []
    errors = json.loads(row["errors"]) if row["errors"] else []

    # Separate empty invoices (no meaningful data) and treat them as errors
    meaningful = []
    for inv in results:
        # Check critical fields: invoice_number, vendor_name, total_amount_due
        if inv.get("invoice_number") or inv.get("vendor_name") or inv.get("total_amount_due"):
            meaningful.append(inv)
        else:
            # Add to errors list with filename if available
            filename = inv.get("filename", "unknown")
            errors.append({
                "filename": filename,
                "error": "No invoice data extracted – file may not be an invoice"
            })

    # Split meaningful invoices by confidence
    approved = [inv for inv in meaningful if inv.get("confidence", 0) >= 0.75]
    review = [inv for inv in meaningful if inv.get("confidence", 0) < 0.75]

    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Headers for invoice sheets
    headers = [
        "invoice_number", "vendor_name", "invoice_date",
        "total_amount_due", "currency", "category", "confidence"
    ]

    # Sheet 1: Approuvé (high confidence)
    ws_approved = wb.create_sheet("Approuvé")
    ws_approved.append(headers)
    for inv in approved:
        ws_approved.append([
            inv.get("invoice_number", ""),
            inv.get("vendor_name", ""),
            inv.get("invoice_date", ""),
            inv.get("total_amount_due", ""),
            inv.get("currency", ""),
            inv.get("category", ""),
            inv.get("confidence", "")
        ])

    # Sheet 2: À revoir (low/medium confidence)
    ws_review = wb.create_sheet("À revoir")
    ws_review.append(headers)
    for inv in review:
        ws_review.append([
            inv.get("invoice_number", ""),
            inv.get("vendor_name", ""),
            inv.get("invoice_date", ""),
            inv.get("total_amount_due", ""),
            inv.get("currency", ""),
            inv.get("category", ""),
            inv.get("confidence", "")
        ])

    # Sheet 3: Erreurs (failed extractions + empty invoices + DB errors)
    ws_errors = wb.create_sheet("Erreurs")
    ws_errors.append(["filename", "error"])
    if errors:
        for err in errors:
            ws_errors.append([err.get("filename", ""), err.get("error", "")])
    else:
        ws_errors.append(["No errors", ""])

    # Save and return
    output_path = OUTPUT_DIR / f"{job_id}.xlsx"
    wb.save(output_path)

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"invoices_{job_id}.xlsx"
    )